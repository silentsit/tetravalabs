#!/usr/bin/env python3
"""Import wide-format Price List USD workbook into product_catalog_usd.json."""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[3]
XLSX_DEFAULT = ROOT / "Price List USD 26_06_26.xlsx"
CATALOG_PATH = ROOT / "product_catalog_usd.json"
TIERED_PATH = ROOT / "packages" / "catalog" / "data" / "tiered-catalog.json"
REVAMP_PACK_PATH = ROOT / "revamp" / "app" / "src" / "data" / "pack-pricing.ts"

STOREFRONT_BY_SOURCE = {
    "Supplies & Reconstitution": "Lab Supplies",
    "GLP-1 / Incretin": "GLP-1 Research",
    "BPC-157 / TB500": "Tissue Repair",
    "Blends": "Research Blends",
    "CJC / Ipamorelin / GHRP": "Growth Hormone Axis",
    "Growth Hormone Axis": "Growth Hormone Axis",
    "Mitochondrial / Metabolic Other": "Metabolic & Mitochondrial",
    "Cosmetic / Copper / Tanning": "Tissue Repair",
    "Longevity / Thymic / Neuropeptides": "Longevity & Neuropeptides",
    "Vitamins & Injectables": "Metabolic & Mitochondrial",
    "Legacy Catalog": "Longevity & Neuropeptides",
}


def slugify(value: str) -> str:
    return re.sub(r"^-+|-+$", "", re.sub(r"[^a-z0-9]+", "-", value.lower()))


def normalize_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def parse_strength(full_name: str) -> tuple[str, str]:
    name = full_name.strip()
    match = re.search(
        r"(\d+(?:\.\d+)?\s*(?:mg|ml|iu|mcg)(?:\s*\([^)]+\))?(?:\s*count(?:\s*\([^)]+\))?)?)",
        name,
        re.I,
    )
    if not match:
        return name, "Standard"
    strength = match.group(1).strip()
    base = re.sub(r"\s+" + re.escape(strength) + r"$", "", name, flags=re.I).strip()
    return base, strength


def _detect_pack_columns(ws) -> dict[str, int]:
    """
    Support both workbook layouts:
    - Legacy: ref, 5tot, 5per, 5save, 10tot, 10per, 10save, 20tot, 20per, 20save
    - 2026-07: ref, 5tot, 5sgd, 5per, 5save, 10tot, 10per, 10save, 20tot, 20per, 20save
    """
    headers = {
        c: str(ws.cell(3, c).value or "").lower().replace("\n", " ")
        for c in range(1, 16)
    }

    def find_col(*needles: str) -> int | None:
        for col, text in headers.items():
            if all(n in text for n in needles):
                return col
        return None

    ref = find_col("1-vial") or find_col("ref") or 4
    tot5 = find_col("5-vial", "total ($)") or find_col("5-vial", "total") or 5
    per5 = find_col("5-vial", "per unit") or 6
    save5 = find_col("5-vial", "savings") or 7
    tot10 = find_col("10-vial", "total") or 8
    per10 = find_col("10-vial", "per unit") or 9
    save10 = find_col("10-vial", "savings") or 10
    tot20 = find_col("20-vial", "total") or 11
    per20 = find_col("20-vial", "per unit") or 12
    save20 = find_col("20-vial", "savings") or 13

    # If a SGD total column sits between 5-total and 5-per, header finder still wins.
    return {
        "ref": ref,
        "tot5": tot5,
        "per5": per5,
        "save5": save5,
        "tot10": tot10,
        "per10": per10,
        "save10": save10,
        "tot20": tot20,
        "per20": per20,
        "save20": save20,
    }


def _num(value, *, required: bool = True) -> float | None:
    if value is None or value == "":
        if required:
            raise ValueError("missing numeric cell")
        return None
    return float(value)


def load_workbook_rows(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Price List USD"]
    cols = _detect_pack_columns(ws)
    rows: list[dict] = []
    current_category = ""

    for row_idx in range(4, ws.max_row + 1):
        category_cell = ws.cell(row_idx, 2).value
        product_cell = ws.cell(row_idx, 3).value
        if category_cell and not product_cell:
            current_category = str(category_cell).strip()
            continue
        if not product_cell:
            continue

        product_name = str(product_cell).strip()
        try:
            pack_tiers = [
                {
                    "tier": "5 vials",
                    "qty": 5,
                    "price_usd": round(_num(ws.cell(row_idx, cols["tot5"]).value), 2),
                    "per_unit_usd": round(_num(ws.cell(row_idx, cols["per5"]).value), 2),
                    "savings_pct": round(_num(ws.cell(row_idx, cols["save5"]).value, required=False) or 0, 4),
                },
                {
                    "tier": "10 vials",
                    "qty": 10,
                    "price_usd": round(_num(ws.cell(row_idx, cols["tot10"]).value), 2),
                    "per_unit_usd": round(_num(ws.cell(row_idx, cols["per10"]).value), 2),
                    "savings_pct": round(_num(ws.cell(row_idx, cols["save10"]).value, required=False) or 0, 4),
                },
                {
                    "tier": "20 vials",
                    "qty": 20,
                    "price_usd": round(_num(ws.cell(row_idx, cols["tot20"]).value), 2),
                    "per_unit_usd": round(_num(ws.cell(row_idx, cols["per20"]).value), 2),
                    "savings_pct": round(_num(ws.cell(row_idx, cols["save20"]).value, required=False) or 0, 4),
                },
            ]
        except (TypeError, ValueError):
            # Skip rows without full pack pricing (e.g. nasal sprays).
            continue

        ref = _num(ws.cell(row_idx, cols["ref"]).value, required=False)
        rows.append(
            {
                "category": current_category,
                "product_name": product_name,
                "slug": slugify(product_name),
                "ref_price_usd": ref,
                "pack_tiers": pack_tiers,
            }
        )
    return rows


def index_old_rows(old_rows: list[dict]) -> dict[str, dict]:
    indexed: dict[str, dict] = {}
    for row in old_rows:
        keys = {
            normalize_key(row["slug"]),
            normalize_key(f"{row['name']} {row['strength']}"),
            normalize_key(row["name"]),
        }
        for key in keys:
            indexed.setdefault(key, row)
    return indexed


def match_old_row(product_name: str, slug: str, indexed: dict[str, dict]) -> dict | None:
    candidates = [
        normalize_key(slug),
        normalize_key(product_name),
    ]
    base, strength = parse_strength(product_name)
    candidates.append(normalize_key(f"{base} {strength}"))
    candidates.append(normalize_key(base))

    for key in candidates:
        if key in indexed:
            return indexed[key]
    return None


def write_revamp_pack_pricing(flat_rows: list[dict]) -> None:
    lines = [
        "/** Auto-generated from Price List USD import. Do not edit manually. */",
        "export type PackTier = {",
        "  tier: string;",
        "  qty: number;",
        "  price: number;",
        "  perUnit: number;",
        "  savingsPct: number;",
        "};",
        "",
        "export const packPricingBySlug: Record<string, PackTier[]> = {",
    ]
    for row in flat_rows:
        tier_objects = []
        for tier in row["pack_tiers"]:
            tier_objects.append(
                "{ tier: "
                + json.dumps(tier["tier"])
                + f", qty: {tier['qty']}, price: {tier['price_usd']}, perUnit: {tier['per_unit_usd']}, savingsPct: {tier['savings_pct']} }}"
            )
        tiers = ",\n      ".join(tier_objects)
        lines.append(f"  {json.dumps(row['slug'])}: [\n      {tiers},\n    ],")
    lines.extend(
        [
            "};",
            "",
            "export function getPackTiers(slug: string): PackTier[] | undefined {",
            "  if (packPricingBySlug[slug]) return packPricingBySlug[slug]",
            "  const normalized = slug.toLowerCase().replace(/[^a-z0-9]/g, '');",
            "  for (const [catalogSlug, tiers] of Object.entries(packPricingBySlug)) {",
            "    if (catalogSlug.replace(/[^a-z0-9]/g, '') === normalized) return tiers",
            "  }",
            "  return undefined",
            "}",
            "",
            "export function getDefaultPackTier(slug: string): PackTier | undefined {",
            "  const tiers = getPackTiers(slug);",
            "  return tiers?.[0];",
            "}",
            "",
        ]
    )
    REVAMP_PACK_PATH.parent.mkdir(parents=True, exist_ok=True)
    REVAMP_PACK_PATH.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else XLSX_DEFAULT
    if not xlsx_path.exists():
        print(f"Missing workbook: {xlsx_path}", file=sys.stderr)
        return 1

    old_rows = json.loads(CATALOG_PATH.read_text(encoding="utf-8-sig"))
    indexed = index_old_rows(old_rows)
    imported = load_workbook_rows(xlsx_path)

    flat_rows: list[dict] = []
    tiered_products: list[dict] = []
    slug_changes: list[tuple[str, str]] = []
    new_products: list[str] = []

    for item in imported:
        old = match_old_row(item["product_name"], item["slug"], indexed)
        base_name, strength = parse_strength(item["product_name"])
        category = item["category"]
        storefront_category = STOREFRONT_BY_SOURCE.get(category, "Growth Factors")

        if old:
            base_name = old["name"]
            strength = old["strength"]
            category = old["category"]
            storefront_category = old.get("storefront_category") or storefront_category
            if old["slug"] != item["slug"]:
                slug_changes.append((old["slug"], item["slug"]))
        else:
            new_products.append(item["product_name"])

        flat_rows.append(
            {
                "category": category,
                "name": base_name,
                "strength": strength,
                "price_usd": item["pack_tiers"][0]["price_usd"],
                "slug": item["slug"],
                "storefront_category": storefront_category,
                "ref_price_usd": item["ref_price_usd"],
                "pack_tiers": item["pack_tiers"],
            }
        )
        tiered_products.append(
            {
                "category": category,
                "storefront_category": storefront_category,
                "name": item["product_name"],
                "slug": item["slug"],
                "ref_price_usd": item["ref_price_usd"],
                "pack_tiers": item["pack_tiers"],
            }
        )

    CATALOG_PATH.write_text(json.dumps(flat_rows, indent=2), encoding="utf-8")
    TIERED_PATH.parent.mkdir(parents=True, exist_ok=True)
    TIERED_PATH.write_text(
        json.dumps({"source": str(xlsx_path), "products": tiered_products}, indent=2),
        encoding="utf-8",
    )
    write_revamp_pack_pricing(flat_rows)

    print(f"Imported {len(flat_rows)} products from {xlsx_path}")
    print(f"Updated {CATALOG_PATH}")
    print(f"Wrote {TIERED_PATH}")
    if slug_changes:
        print(f"Slug updates ({len(slug_changes)}):")
        for old_slug, new_slug in slug_changes:
            print(f"  {old_slug} -> {new_slug}")
    if new_products:
        print(f"New products ({len(new_products)}): {', '.join(new_products)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
