from copy import copy
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl


SRC = Path(r"C:\Users\user\Downloads\Tiered Pricing-upload.cleaned.xlsx")
OUT = Path(r"C:\Users\user\Downloads\Tiered Pricing-upload.cleaned-with-new-items.xlsx")
CATEGORY = "Longevity / Thymic / Neuropeptides"
NEW_PRODUCTS = [
    ("Pinealon 10mg", 75.00),
    ("Humanin 10mg", 129.00),
    ("Adamax 10mg", 79.00),
    ("Dihexa 10mg", 99.00),
    ("Cerebrolysin 10mg", 69.00),
]


def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)


def money(value):
    return float(Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def savings(base, per):
    return float((Decimal("1") - Decimal(str(per)) / Decimal(str(base))).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP))


def total_ok(per, qty):
    total = per * Decimal(qty)
    return total == total.to_integral_value() and int(total) % 10 in (0, 5, 9)


def allowed_per_vial_prices(base, qty):
    upper = Decimal(str(base))
    candidates = set()
    max_int = int(upper.to_integral_value(rounding=ROUND_HALF_UP)) + 5

    if qty == 5:
        for n in range(1, max_int + 1):
            if n % 10 in (0, 5, 9):
                candidates.add(Decimal(n))
    elif qty == 10:
        for n in range(0, max_int + 1):
            candidates.add(Decimal(n) + Decimal("0.90"))
            candidates.add(Decimal(n) + Decimal("0.50"))
        for n in range(1, max_int + 1):
            if n % 10 in (0, 5, 9):
                candidates.add(Decimal(n))
    elif qty == 20:
        for n in range(0, max_int + 1):
            candidates.add(Decimal(n) + Decimal("0.50"))
        for n in range(1, max_int + 1):
            if n % 10 in (0, 5, 9):
                candidates.add(Decimal(n))

    return sorted(p for p in candidates if p > 0)


def choose_price(base, qty, previous=None):
    base_d = Decimal(str(base))
    target_factor = {5: Decimal("0.86"), 10: Decimal("0.80"), 20: Decimal("0.74")}[qty]
    target = base_d * target_factor
    max_allowed = base_d * Decimal("0.95")
    if previous is not None:
        max_allowed = min(max_allowed, Decimal(str(previous)) - Decimal("0.10"))

    candidates = [p for p in allowed_per_vial_prices(base, qty) if p <= max_allowed and total_ok(p, qty)]
    if not candidates:
        candidates = [p for p in allowed_per_vial_prices(base, qty) if p <= base_d and total_ok(p, qty)]
    if previous is not None:
        below_previous = [p for p in candidates if p <= Decimal(str(previous)) - Decimal("0.10")]
        if below_previous:
            candidates = below_previous

    selected = min(candidates, key=lambda p: (abs(p - target), p))
    return money(selected)


def copy_row_style(ws, src_row, dst_row):
    if ws.row_dimensions[src_row].height:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(1, 11):
        copy_cell_style(ws.cell(src_row, col), ws.cell(dst_row, col))


def write_tier(ws, row, style_row, category, product, tier, qty, total, per, save, ref, note):
    copy_row_style(ws, style_row, row)
    values = {
        2: category,
        3: product,
        4: tier,
        5: qty,
        6: total,
        7: per,
        8: save,
        9: ref,
        10: note,
    }
    for col, value in values.items():
        ws.cell(row, col).value = value
    ws.cell(row, 5).number_format = "0"
    ws.cell(row, 6).number_format = "$#,##0"
    ws.cell(row, 7).number_format = "$#,##0.00"
    ws.cell(row, 8).number_format = "0.0%"
    ws.cell(row, 9).number_format = "$#,##0.00"


def main():
    wb = openpyxl.load_workbook(SRC)
    ws = wb["Tiered Pricing"]

    existing = {ws.cell(r, 3).value for r in range(1, ws.max_row + 1)}
    products_to_add = [(name, price) for name, price in NEW_PRODUCTS if name not in existing]
    if not products_to_add:
        wb.save(OUT)
        print("No new products to add.")
        return

    insert_at = None
    for row in range(1, ws.max_row + 1):
        category = ws.cell(row, 2).value
        product = ws.cell(row, 3).value
        if product and category == CATEGORY:
            insert_at = row + 4

    if insert_at is None:
        raise RuntimeError(f"Category not found: {CATEGORY}")

    style_rows = {
        1: insert_at - 4,
        5: insert_at - 3,
        10: insert_at - 2,
        20: insert_at - 1,
    }
    rows_needed = len(products_to_add) * 4
    ws.insert_rows(insert_at, rows_needed)

    current = insert_at
    for product, base in products_to_add:
        base = money(base)
        p5 = choose_price(base, 5)
        p10 = choose_price(base, 10, p5)
        p20 = choose_price(base, 20, p10)
        tiers = [
            (style_rows[1], "1 Vial", 1, base, base, 0, base, "Reference only"),
            (style_rows[5], "5 Vials", 5, money(p5 * 5), p5, savings(base, p5), base, "Offer tier"),
            (style_rows[10], "10 Vials", 10, money(p10 * 10), p10, savings(base, p10), base, "Offer tier"),
            (style_rows[20], "20 Vials", 20, money(p20 * 20), p20, savings(base, p20), base, "Offer tier"),
        ]
        for style_row, tier, qty, total, per, save, ref, note in tiers:
            write_tier(ws, current, style_row, CATEGORY, product, tier, qty, total, per, save, ref, note)
            current += 1
        print(f"{product}: 1=${base:.2f}; 5=${p5:.2f}/vial (${p5 * 5:.0f}); 10=${p10:.2f}/vial (${p10 * 10:.0f}); 20=${p20:.2f}/vial (${p20 * 20:.0f})")

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
