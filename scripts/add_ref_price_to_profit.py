from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

PRICING_PATH = Path(r"C:\Users\user\Downloads\Tiered Pricing-upload.cleaned-with-new-items.xlsx")
ANALYSIS_PATH = Path(r"C:\Users\user\Downloads\Tiered Pricing Profit Analysis.xlsx")


def clone_style(src, dst):
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def load_reference_prices():
    wb = openpyxl.load_workbook(PRICING_PATH, data_only=True)
    ws = wb["Tiered Pricing"]
    refs = {}
    for row in range(1, ws.max_row + 1):
        product = ws.cell(row, 3).value
        qty = ws.cell(row, 5).value
        per_vial = ws.cell(row, 7).value
        if product and qty == 1 and per_vial is not None:
            refs[product] = float(per_vial)
    return refs


def main():
    refs = load_reference_prices()
    wb = openpyxl.load_workbook(ANALYSIS_PATH)
    ws = wb["Profit Analysis"]

    insert_col = 3
    ws.insert_cols(insert_col)
    ws.cell(1, insert_col).value = "1-Vial Sell USD (Ref)"
    clone_style(ws.cell(1, 2), ws.cell(1, insert_col))
    ws.cell(1, insert_col).font = Font(bold=True)
    ws.cell(1, insert_col).fill = PatternFill("solid", fgColor="D9EAF7")

    for row in range(2, ws.max_row + 1):
        product = ws.cell(row, 2).value
        ref = refs.get(product)
        cell = ws.cell(row, insert_col)
        cell.value = ref
        cell.number_format = "$#,##0.00"
        clone_style(ws.cell(row, 2), cell)

    ws.column_dimensions[openpyxl.utils.get_column_letter(insert_col)].width = 22
    ws.auto_filter.ref = ws.dimensions
    wb.save(ANALYSIS_PATH)
    print(f"Updated {ANALYSIS_PATH}")
    print(f"Reference prices added for {sum(1 for r in range(2, ws.max_row + 1) if ws.cell(r, insert_col).value is not None)} rows")


if __name__ == "__main__":
    main()
