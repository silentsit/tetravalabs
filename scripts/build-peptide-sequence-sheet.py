#!/usr/bin/env python3
"""Build peptide Excel sheet (name, chemical formula, concentration, vial, color) from Price List USD.xlsx."""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PRICE_LIST = Path(r"c:\Users\user\Downloads\Price List USD.xlsx")
OUTPUT = Path(r"c:\Users\user\Downloads\Peptide Sequence Sheet.xlsx")
ENRICHMENT = Path(__file__).resolve().parents[1] / "packages" / "catalog" / "data" / "product-enrichment.json"

# Empirical molecular formulas (supplier COA / PubChem where available).
FORMULAS: dict[str, str] = {
    "5-amino-1mq": "C10H11N3",
    "acetic acid water": "C2H4O2 (in H2O)",
    "bacteriostatic water": "H2O",
    "benzyl alcohol": "C7H8O (in H2O)",
    "lemon bottle": "Multi-component solution (no single formula)",
    "lipo-c": "Multi-component lipotropic blend (no single formula)",
    "bpc-157": "C62H98N16O22",
    "tb-500": "C212H350N56O78S",
    "tb500": "C212H350N56O78S",
    "ghk-cu": "C14H24CuN6O4",
    "ghk cu": "C14H24CuN6O4",
    "cu": "C14H24CuN6O4",
    "kpv": "C16H30N4O4",
    "ll-37": "C205H340N60O53",
    "aod-9604": "C78H123N21O23S",
    "cagrilintide": "C194H312N54O59",
    "semaglutide": "C187H291N45O59",
    "tirzepatide": "C225H348N48O68",
    "retatrutide": "C267H402N64O78",
    "mazdutide": "C225H348N48O68",
    "survodutide": "C187H281N45O59",
    "ipamorelin": "C38H49N9O5",
    "cjc-1295 with dac": "C165H269N47O46",
    "cjc-1295 without dac": "C152H252N44O42",
    "cjc-1295": "C152H252N44O42",
    "sermorelin": "C149H244N44O42",
    "tesamorelin": "C221H366N72O67S",
    "ghrp-2": "C45H56N12O6",
    "ghrp-6": "C46H56N12O6",
    "hexarelin": "C47H58N12O6",
    "igf-1 lr3": "C473H711N131O101S10",
    "igf-1 des": "C119H190N26O29S",
    "mgf": "PEGylated splice variant (no single formula)",
    "peg-mgf": "PEGylated splice variant (no single formula)",
    "epithalon": "C14H22N4O9",
    "epitalon": "C14H22N4O9",
    "selank": "C33H57N11O9",
    "semax": "C35H52N12O9",
    "dsip": "C35H48N10O15",
    "oxytocin": "C43H66N12O12S2",
    "melanotan 1": "C78H111N21O19",
    "melanotan 2": "C50H69N15O9",
    "melanotan ii": "C50H69N15O9",
    "bremelanotide": "C50H68N14O10",
    "pt-141": "C50H68N14O10",
    "hcg": "Glycoprotein hormone (no single empirical formula)",
    "hmg": "Glycoprotein hormone (no single empirical formula)",
    "hgh 191aa": "C990H1528N262O300S7",
    "hgh": "C990H1528N262O300S7",
    "mk-677": "C27H36N4O5",
    "nad+": "C21H27N7O14P2",
    "nad": "C21H27N7O14P2",
    "glutathione": "C10H17N3O6S",
    "l-carnitine": "C7H15NO3",
    "l-glu": "C5H9NO4",
    "vitamin b-12": "C63H88CoN14O14P",
    "b-12": "C63H88CoN14O14P",
    "aicar": "C9H14N4O5",
    "adipotide": "C152H252N44O42",
    "ara-290": "C23H40N8O9",
    "foxo4-dri": "Proprietary D-retro-inverso peptide (verify COA)",
    "snap-8": "C41H62N10O15",
    "mots-c": "C101H168N30O21",
    "ss-31": "C32H49N9O5",
    "thymosin alpha-1": "C129H215N33O55",
    "thymosin alpha 1": "C129H215N33O55",
    "thymalin": "Peptide complex (no single formula)",
    "thymalin complex": "Peptide complex (no single formula)",
    "vip": "C147H237N43O43S",
    "gonadorelin": "C55H75N17O13",
    "kisspeptin-10": "C63H83N17O14",
    "kisspeptin 10": "C63H83N17O14",
    "pinealon": "C14H22N4O9",
    "pinealon capsules": "C14H22N4O9",
    "bpc-157 capsules": "C62H98N16O22",
    "dermorphin": "C26H35N5O5",
    "copper repair blend": "C14H24CuN6O4 | C212H350N56O78S | C62H98N16O22 | C16H30N4O4",
    "glow blend": "C62H98N16O22 | C212H350N56O78S | C14H24CuN6O4",
    "glow enhanced": "C212H350N56O78S | C62H98N16O22 | C14H24CuN6O4",
    "glow bpc-157 + tb500 + ghk-cu": "C62H98N16O22 | C212H350N56O78S | C14H24CuN6O4",
    "cagrilintide + semaglutide": "C194H312N54O59 | C187H291N45O59",
    "cjc-1295 ipamorelin blend": "C152H252N44O42 | C38H49N9O5",
    "cjc-1295 sermorelin ipamorelin blend": "C149H244N44O42 | C38H49N9O5",
    "bpc-157 + tb500": "C62H98N16O22 | C212H350N56O78S",
    "bpc-157 tb500": "C62H98N16O22 | C212H350N56O78S",
    "cu + tb500 + bpc-157 + kpv": "C14H24CuN6O4 | C212H350N56O78S | C62H98N16O22 | C16H30N4O4",
    "adamax": "C37H54N12O10",
    "cerebrolysin": "Multi-peptide mixture (no single formula)",
    "dihexa": "C27H32N6O5",
    "humanin": "C119H204N36O32",
    "pinealon peptide": "C14H22N4O9",
}

DEFAULT_VIAL = "10ml"

SOURCE_CATEGORY_MAP = {
    "Supplies & Reconstitution": "Lab Supplies",
    "BPC-157 / TB500": "Growth Factors",
    "CJC / Ipamorelin / GHRP": "Growth Factors",
    "Growth Hormone Axis": "Growth Factors",
    "Mitochondrial / Metabolic Other": "Growth Factors",
    "Cosmetic / Copper / Tanning": "Growth Factors",
    "Longevity / Thymic / Neuropeptides": "Growth Factors",
    "Vitamins & Injectables": "Growth Factors",
    "Legacy Catalog": "Growth Factors",
    "Blends": "Research Blends",
}

BLEND_PRODUCTS = {
    "BPC-157 5mg + TB500 5mg",
    "CU 50mg + TB500 10mg + BPC-157 10mg + KPV 10mg",
    "Glow BPC-157 + TB500 + GHK-Cu",
    "Glow TB500 10mg + BPC-157 10mg + GHK-Cu 50mg",
    "CJC-1295 without DAC / Ipamorelin Blend",
    "CJC-1295 without DAC / Sermorelin / Ipamorelin Blend",
    "Cagrilintide + Semaglutide",
}

GROWTH_FROM_GLP1 = {"Sermorelin", "Tesamorelin"}

GLP1_PRODUCTS = {
    "Semaglutide",
    "Tirzepatide",
    "Retatrutide",
    "Cagrilintide",
    "Mazdutide",
    "Survodutide",
    "5-Amino-1MQ",
    "5-amino-1mq",
    "AOD-9604",
}

CATEGORY_COLORS = {
    "GLP-1 Research": "#3357FF",
    "Growth Factors": "#33FF57",
    "Research Blends": "#FF5733",
    "Lab Supplies": "#FFC133",
}


def load_enrichment_formulas() -> dict[str, str]:
    if not ENRICHMENT.exists():
        return {}
    data = json.loads(ENRICHMENT.read_text(encoding="utf-8"))
    out: dict[str, str] = {}
    for name, meta in data.items():
        formula = meta.get("molecular_formula")
        if formula:
            out[normalize_key(name)] = formula
    return out


def normalize_key(name: str) -> str:
    text = name.lower().strip()
    text = re.sub(r"\([^)]*\)", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def strip_strength(text: str) -> str:
    return re.sub(r"\s*\d+(?:\.\d+)?\s*(?:mg|mcg|iu)\b", "", text, flags=re.I).strip()


def strip_blend_total(name: str) -> str:
    return re.sub(r"\s*\(\d+(?:\.\d+)?\s*(?:mg|mcg|iu)\)\s*$", "", name, flags=re.I).strip()


def parse_base_name(full_name: str) -> str:
    name = strip_blend_total(full_name.strip())
    strength = re.search(
        r"\s+(\d+(?:\.\d+)?\s*(?:mg|mcg|iu)(?:\s*\([^)]+\))?|(?:\d+\s*count(?:\s*\([^)]+\))?))\s*$",
        name,
        re.I,
    )
    if strength:
        name = name[: strength.start()].strip()
    return name


def resolve_storefront_category(full_name: str, source_category: str) -> str:
    blend_key = strip_blend_total(full_name.strip())
    if blend_key in BLEND_PRODUCTS:
        return "Research Blends"

    base_name = parse_base_name(full_name)
    if base_name in BLEND_PRODUCTS:
        return "Research Blends"

    if source_category == "GLP-1 / Incretin":
        if base_name in GROWTH_FROM_GLP1:
            return "Growth Factors"
        if base_name in GLP1_PRODUCTS or normalize_key(base_name) in {normalize_key(x) for x in GLP1_PRODUCTS}:
            return "GLP-1 Research"
        return "Growth Factors"

    return SOURCE_CATEGORY_MAP.get(source_category, "Growth Factors")


def color_for_category(category: str) -> str:
    return CATEGORY_COLORS.get(category, "#FFFFFF")


def parse_product(full_name: str) -> tuple[str, str, str]:
    name = full_name.strip()
    concentration = "—"
    vial_size = DEFAULT_VIAL

    volume = re.search(r"(\d+(?:\.\d+)?\s*ml)\b", name, re.I)
    count = re.search(r"(\d+\s*count)\b", name, re.I)
    if volume:
        vial_size = volume.group(1).lower().replace(" ", "")
    elif count:
        vial_size = count.group(1)

    total_blend = re.search(r"\((\d+(?:\.\d+)?\s*(?:mg|mcg|iu))\)\s*$", name, re.I)
    if total_blend:
        concentration = total_blend.group(1).strip()
        name = name[: total_blend.start()].strip()
    else:
        strength = re.search(
            r"(\d+(?:\.\d+)?\s*(?:mg|mcg|iu)(?:\s*\([^)]+\))?)\s*$",
            name,
            re.I,
        )
        if strength:
            concentration = strength.group(1).strip()
            name = name[: strength.start()].strip()

    if count:
        name = re.sub(r"\s*\(\s*\d+\s*count\s*\)", "", name, flags=re.I).strip()
        name = re.sub(r"\s+\d+\s*count\b", "", name, flags=re.I).strip()

    if "+" in name:
        parts = []
        for part in re.split(r"\s*\+\s*", name):
            cleaned = strip_strength(part)
            cleaned = cleaned.replace("TB500", "TB-500")
            if cleaned:
                parts.append(cleaned)
        name = " + ".join(parts)
    else:
        name = name.replace("TB500", "TB-500")

    return name, concentration, vial_size


def lookup_formula_single(peptide_name: str, enrichment: dict[str, str]) -> str:
    key = normalize_key(peptide_name)

    ordered = sorted(FORMULAS.items(), key=lambda item: len(item[0]), reverse=True)
    for pattern, formula in ordered:
        if key == pattern or pattern in key:
            return formula

    for pattern, formula in sorted(enrichment.items(), key=lambda item: len(item[0]), reverse=True):
        if key == pattern or pattern in key:
            return formula

    return "Formula not on file — verify against product COA"


def lookup_formula(peptide_name: str, enrichment: dict[str, str]) -> str:
    if "+" in peptide_name:
        parts = [lookup_formula_single(part.strip(), enrichment) for part in peptide_name.split("+")]
        parts = [formula for formula in parts if formula]
        if parts:
            return " | ".join(dict.fromkeys(parts))

    return lookup_formula_single(peptide_name, enrichment)


_SUBSCRIPT = str.maketrans("0123456789", "₀₁₂₃₄₅₆₇₈₉")
_SKIP_SUBSCRIPT_MARKERS = (
    "no single",
    "multi-",
    "glycoprotein",
    "proprietary",
    "pegylated",
    "peptide complex",
    "verify coa",
    "not on file",
    "formula not",
)


def format_formula_subscript(text: str) -> str:
    """Render empirical formulas with subscript counts, e.g. C62H98N16O22 -> C₆₂H₉₈N₁₆O₂₂."""

    if not text:
        return text

    value = str(text).strip()
    lower = value.lower()
    if any(marker in lower for marker in _SKIP_SUBSCRIPT_MARKERS):
        return value

    def convert_segment(segment: str) -> str:
        out: list[str] = []
        index = 0
        while index < len(segment):
            match = re.match(r"([A-Z][a-z]?)", segment[index:])
            if match:
                out.append(match.group(1))
                index += len(match.group(1))
                digits = re.match(r"(\d+)", segment[index:])
                if digits:
                    out.append(digits.group(1).translate(_SUBSCRIPT))
                    index += len(digits.group(1))
            else:
                out.append(segment[index])
                index += 1
        return "".join(out)

    return " | ".join(convert_segment(part.strip()) for part in value.split("|"))


_SUBSCRIPT_DIGITS = set("0123456789" + "₀₁₂₃₄₅₆₇₈₉")
_FORMULA_BASE_COLOR = "FFFFFFFF"
_FORMULA_DIGIT_COLOR = "FFFF0000"


def _is_formula_digit(char: str) -> bool:
    return char in _SUBSCRIPT_DIGITS


def formula_rich_text(text: str) -> CellRichText:
    """Return a formula with subscript counts and red digit styling."""

    formatted = format_formula_subscript(text)
    if not formatted:
        return CellRichText()

    blocks: list[TextBlock] = []
    chunk = ""
    chunk_is_digit = _is_formula_digit(formatted[0])

    for char in formatted:
        flag = _is_formula_digit(char)
        if chunk and flag != chunk_is_digit:
            color = _FORMULA_DIGIT_COLOR if chunk_is_digit else _FORMULA_BASE_COLOR
            blocks.append(TextBlock(InlineFont(color=color), chunk))
            chunk = ""
            chunk_is_digit = flag
        chunk += char
        chunk_is_digit = flag

    if chunk:
        color = _FORMULA_DIGIT_COLOR if chunk_is_digit else _FORMULA_BASE_COLOR
        blocks.append(TextBlock(InlineFont(color=color), chunk))

    return CellRichText(*blocks)


def load_products() -> list[dict]:
    enrichment = load_enrichment_formulas()
    wb = openpyxl.load_workbook(PRICE_LIST, read_only=True, data_only=True)
    ws = wb["Tiered Price List"]
    rows: list[dict] = []
    current_category = ""
    for row in ws.iter_rows(min_row=5, values_only=True):
        category_cell = row[1]
        product_name = row[2]
        if category_cell and not product_name:
            current_category = str(category_cell).strip()
            continue
        if not product_name or not str(product_name).strip():
            continue
        full = str(product_name).strip()
        peptide_name, concentration, vial_size = parse_product(full)
        storefront_category = resolve_storefront_category(full, current_category)
        rows.append(
            {
                "peptide_name": peptide_name,
                "chemical_formula": lookup_formula(peptide_name, enrichment),
                "concentration": concentration,
                "vial_size": vial_size,
                "category": storefront_category,
                "color_code": color_for_category(storefront_category),
            }
        )
    wb.close()
    return rows


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(color="374151", style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Peptide Name", "Chemical Formula", "Concentration", "Vial Size", "Color Code"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border

    widths = [28, 36, 16, 12, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A2"


def write_excel(rows: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Peptide Formulas"
    style_sheet(ws)

    thin = Side(color="374151", style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    body_font = Font(color="FFFFFF")
    row_fill_a = PatternFill("solid", fgColor="111827")
    row_fill_b = PatternFill("solid", fgColor="1F2937")

    for i, row in enumerate(rows, start=2):
        values = [
            row["peptide_name"],
            formula_rich_text(row["chemical_formula"]),
            row["concentration"],
            row["vial_size"],
            row["color_code"],
        ]
        fill = row_fill_a if i % 2 == 0 else row_fill_b
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=value)
            cell.fill = fill
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    wb.save(OUTPUT)


def main() -> None:
    rows = load_products()
    write_excel(rows)
    print(f"Wrote {len(rows)} rows to {OUTPUT}")


if __name__ == "__main__":
    main()
