from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

COST_PATH = Path(r"C:\Users\user\Downloads\Peptide Cost Sheet.tallied-corrected.xlsx")
PRICING_PATH = Path(r"C:\Users\user\Downloads\Tiered Pricing-upload.cleaned-with-new-items.xlsx")
OUT_PATH = Path(r"C:\Users\user\Downloads\Tiered Pricing Profit Analysis.xlsx")

USD_FALLBACK_RATE = 0.14682143


def clone_style(src, dst):
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def last_used_row(ws, col=2, start=2):
    last = 1
    for row in range(start, ws.max_row + 1):
        if ws.cell(row, col).value not in (None, ""):
            last = row
    return last


def copy_costs_sheet(src_wb, dst_wb):
    src = src_wb["Sheet1"]
    dst = dst_wb.create_sheet("Costs")
    end = last_used_row(src, col=2)
    end = max(end, last_used_row(src, col=1))

    for row in range(1, end + 1):
        for col in range(1, 5):
            dst.cell(row, col).value = src.cell(row, col).value
            clone_style(src.cell(row, col), dst.cell(row, col))

    for row in range(3, end + 1):
        if dst.cell(row, 3).value not in (None, ""):
            dst.cell(row, 4).value = (
                f'=IF($C{row}="","",IFERROR(GOOGLEFINANCE("currency:cnyusd")*$C{row},Settings!$B$1*$C{row}))'
            )

    dst.column_dimensions["A"].width = 28
    dst.column_dimensions["B"].width = 42
    dst.column_dimensions["C"].width = 18
    dst.column_dimensions["D"].width = 18
    return end


def copy_pricing_sheet(src_wb, dst_wb):
    src = src_wb["Tiered Pricing"]
    dst = dst_wb.create_sheet("Tiered Pricing")
    end = 1
    for row in range(1, src.max_row + 1):
        if any(src.cell(row, col).value not in (None, "") for col in range(2, 11)):
            end = row
            for col in range(2, 11):
                dst.cell(row, col - 1).value = src.cell(row, col).value
                clone_style(src.cell(row, col), dst.cell(row, col - 1))

    # Helper lookup key: Product|Qty (columns B + D on copied pricing sheet)
    dst.cell(2, 10).value = "Lookup Key"
    clone_style(dst.cell(2, 2), dst.cell(2, 10))
    dst.cell(2, 10).font = Font(bold=True)
    for row in range(3, end + 1):
        if dst.cell(row, 2).value and dst.cell(row, 4).value not in (None, ""):
            dst.cell(row, 10).value = f'=IF($B{row}="","",$B{row}&"|"&$D{row})'

    widths = {1: 28, 2: 42, 3: 12, 4: 8, 5: 14, 6: 12, 7: 16, 8: 12, 9: 16, 10: 42}
    for col, width in widths.items():
        dst.column_dimensions[get_column_letter(col)].width = width
    return end


def load_product_groups(pricing_ws):
    groups = {}
    for row in range(3, pricing_ws.max_row + 1):
        category = pricing_ws.cell(row, 2).value
        product = pricing_ws.cell(row, 3).value
        tier = pricing_ws.cell(row, 4).value
        qty = pricing_ws.cell(row, 5).value
        if not product or qty not in (1, 5, 10, 20):
            continue
        group = groups.setdefault(product, {"category": category, "product": product, "has_ref": False, "tiers": []})
        if category:
            group["category"] = category
        if qty == 1:
            group["has_ref"] = True
        else:
            group["tiers"].append({"tier": tier, "qty": int(qty)})
    return groups


def add_settings_sheet(wb):
    ws = wb.create_sheet("Settings")
    ws["A1"] = "CNY→USD fallback rate"
    ws["B1"] = USD_FALLBACK_RATE
    ws["A3"] = "Notes"
    ws["B3"] = "Costs!D uses GOOGLEFINANCE with this fallback. Profit Analysis recalculates from Costs and Tiered Pricing."
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 48
    ws["B1"].number_format = "0.00000000"


def build_profit_analysis(wb, groups):
    ws = wb.active
    ws.title = "Profit Analysis"

    headers = [
        "Category",
        "Product",
        "Tier",
        "Qty",
        "Sell Per Vial USD",
        "Sell Total USD",
        "Cost Per Vial USD",
        "Profit Per Vial USD",
        "Total Profit USD",
        "Margin %",
        "Status",
        "Cost Sheet Row",
        "Cost RMB",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    ref_fill = PatternFill("solid", fgColor="E8EEF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    row_num = 2
    for product in sorted(groups, key=lambda p: (groups[p]["category"] or "", p)):
        group = groups[product]
        category = group["category"]

        def write_row(tier, qty, fill=None):
            nonlocal row_num
            ws.cell(row_num, 1).value = category
            ws.cell(row_num, 2).value = product
            ws.cell(row_num, 3).value = tier
            ws.cell(row_num, 4).value = qty
            r = row_num

            ws.cell(r, 5).value = (
                f'=IFERROR(XLOOKUP($B{r}&"|"&$D{r},\'Tiered Pricing\'!$J:$J,\'Tiered Pricing\'!$F:$F),"")'
            )
            ws.cell(r, 6).value = f'=IF($E{r}="","",$E{r}*$D{r})'
            ws.cell(r, 7).value = f'=IFERROR(XLOOKUP($B{r},Costs!$B:$B,Costs!$D:$D),"")'
            ws.cell(r, 8).value = f'=IF(OR($E{r}="",$G{r}=""),"",$E{r}-$G{r})'
            ws.cell(r, 9).value = f'=IF($H{r}="","",$H{r}*$D{r})'
            ws.cell(r, 10).value = f'=IF(OR($E{r}="",$H{r}=""),"",$H{r}/$E{r})'
            ws.cell(r, 11).value = (
                f'=IF($G{r}="","MISSING COST",IF($C{r}="1 Vial (Ref)","REFERENCE",IF($H{r}<0,"LOSS","PROFIT")))'
            )
            ws.cell(r, 12).value = f'=IFERROR(MATCH($B{r},Costs!$B:$B,0)+1,"")'
            ws.cell(r, 13).value = f'=IFERROR(XLOOKUP($B{r},Costs!$B:$B,Costs!$C:$C),"")'

            for col in (5, 6, 7, 8, 9):
                ws.cell(r, col).number_format = "$#,##0.00"
            ws.cell(r, 10).number_format = "0.0%"
            ws.cell(r, 13).number_format = "0.00"

            if fill:
                for col in range(1, 14):
                    ws.cell(r, col).fill = fill
            row_num += 1

        if group["has_ref"]:
            write_row("1 Vial (Ref)", 1, ref_fill)
        for tier in sorted(group["tiers"], key=lambda t: t["qty"]):
            write_row(tier["tier"], tier["qty"])

    for col in range(1, 14):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_num - 1}"
    ws.freeze_panes = "A2"


def main():
    cost_wb = openpyxl.load_workbook(COST_PATH, data_only=True)
    pricing_wb = openpyxl.load_workbook(PRICING_PATH, data_only=True)
    groups = load_product_groups(pricing_wb["Tiered Pricing"])

    out_wb = openpyxl.Workbook()
    add_settings_sheet(out_wb)
    copy_costs_sheet(openpyxl.load_workbook(COST_PATH, data_only=False), out_wb)
    copy_pricing_sheet(pricing_wb, out_wb)
    build_profit_analysis(out_wb, groups)

    # Move Profit Analysis to first tab for convenience.
    out_wb.move_sheet("Profit Analysis", offset=-2)

    out_wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    print(f"Products: {len(groups)}")
    print("Sheets:", out_wb.sheetnames)


if __name__ == "__main__":
    main()
