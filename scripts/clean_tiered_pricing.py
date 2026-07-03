from copy import copy
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl


SRC = Path(r"C:\Users\user\Downloads\Tiered Pricing-upload.xlsx")
OUT = Path(r"C:\Users\user\Downloads\Tiered Pricing-upload.cleaned.xlsx")


def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)
    if src_cell.comment:
        dst_cell.comment = copy(src_cell.comment)


def money(value):
    return float(Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def savings(base, per):
    if not base:
        return 0
    return float((Decimal("1") - Decimal(str(per)) / Decimal(str(base))).quantize(Decimal("0.0001"), rounding=ROUND_HALF_UP))


def total_ok(per, qty):
    total = per * Decimal(qty)
    if total != total.to_integral_value():
        return False
    return int(total) % 10 in (0, 5, 9)


def allowed_per_vial_prices(base, qty):
    # Allowed per-vial endings while keeping tier totals as whole numbers ending in 9, 5, or 0.
    upper = Decimal(str(base))
    candidates = set()
    max_int = int(upper.to_integral_value(rounding=ROUND_HALF_UP)) + 5

    if qty == 5:
        # Decimal per-vial prices cannot produce whole-number totals for qty 5.
        for n in range(1, max_int + 1):
            if n % 10 in (0, 5, 9):
                candidates.add(Decimal(n))
    elif qty == 10:
        # .90 => total ends in 9, .50 => total ends in 5.
        for n in range(0, max_int + 1):
            for cents in ("0.90", "0.50"):
                value = Decimal(n) + Decimal(cents)
                if value > 0:
                    candidates.add(value)
        for n in range(1, max_int + 1):
            if n % 10 in (0, 5, 9):
                candidates.add(Decimal(n))
    elif qty == 20:
        # .50 => total ends in 0. .90 would make the total end in 8, so skip it.
        for n in range(0, max_int + 1):
            value = Decimal(n) + Decimal("0.50")
            if value > 0:
                candidates.add(value)
        for n in range(1, max_int + 1):
            if n % 10 in (0, 5, 9):
                candidates.add(Decimal(n))

    return sorted(candidates)


def choose_price(base, qty, previous=None):
    base_d = Decimal(str(base))
    target_factor = {5: Decimal("0.86"), 10: Decimal("0.80"), 20: Decimal("0.74")}[qty]
    target = base_d * target_factor
    max_allowed = base_d * Decimal("0.95")
    if previous is not None:
        max_allowed = min(max_allowed, Decimal(str(previous)) - Decimal("0.10"))

    candidates = [p for p in allowed_per_vial_prices(base, qty) if p <= max_allowed and total_ok(p, qty)]
    if not candidates:
        candidates = [p for p in allowed_per_vial_prices(base, qty) if p <= base_d and total_ok(p, qty)]
    if previous is not None:
        below_previous = [p for p in candidates if p <= Decimal(str(previous)) - Decimal("0.10")]
        if below_previous:
            candidates = below_previous
    if not candidates:
        candidates = [Decimal("0.50")]

    selected = min(candidates, key=lambda p: (abs(p - target), p))
    return money(selected)


def main():
    wb = openpyxl.load_workbook(SRC)
    ws = wb["Tiered Pricing"]

    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = ws.title
    new_ws.freeze_panes = ws.freeze_panes
    new_ws.sheet_view.showGridLines = ws.sheet_view.showGridLines

    for key, dim in ws.column_dimensions.items():
        new_dim = new_ws.column_dimensions[key]
        new_dim.width = dim.width
        new_dim.hidden = dim.hidden
    new_ws.column_dimensions["J"].width = 18

    def copy_row_style(src_row, dst_row):
        if ws.row_dimensions[src_row].height:
            new_ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
        for col in range(1, 11):
            copy_cell_style(ws.cell(src_row, col), new_ws.cell(dst_row, col))

    def write_row_from(src_row, dst_row, values):
        copy_row_style(src_row, dst_row)
        for col, value in values.items():
            new_ws.cell(dst_row, col).value = value

    for row in range(1, 4):
        copy_row_style(row, row)
        for col in range(1, 10):
            new_ws.cell(row, col).value = ws.cell(row, col).value
    new_ws["B2"] = "TIERED PRICING -- 1 / 5 / 10 / 20 VIAL (Shop-Ready Price Endings)"
    new_ws["J3"] = "Pricing Note"
    copy_cell_style(ws["H3"], new_ws["J3"])

    out_row = 4
    row = 4
    changed = []
    product_count = 0

    while row <= ws.max_row:
        category = ws.cell(row, 2).value
        product = ws.cell(row, 3).value
        qty = ws.cell(row, 5).value

        if category and not product:
            write_row_from(row, out_row, {col: ws.cell(row, col).value for col in range(1, 10)})
            try:
                new_ws.merge_cells(start_row=out_row, start_column=2, end_row=out_row, end_column=8)
            except ValueError:
                pass
            out_row += 1
            row += 1
            continue

        if product and qty == 1:
            base = ws.cell(row, 9).value or ws.cell(row, 6).value
            base = money(base)
            product_count += 1

            p5 = choose_price(base, 5)
            p10 = choose_price(base, 10, p5)
            p20 = choose_price(base, 20, p10)

            tiers = [
                (row, "1 Vial", 1, base, base, 0, base, "Reference only"),
                (row + 1, "5 Vials", 5, money(p5 * 5), p5, savings(base, p5), base, "Offer tier"),
                (row + 2, "10 Vials", 10, money(p10 * 10), p10, savings(base, p10), base, "Offer tier"),
                (row + 2, "20 Vials", 20, money(p20 * 20), p20, savings(base, p20), base, "Offer tier"),
            ]
            for src_style_row, tier, q, total, per, save, ref, note in tiers:
                write_row_from(src_style_row, out_row, {
                    1: ws.cell(row, 1).value,
                    2: category,
                    3: product,
                    4: tier,
                    5: q,
                    6: total,
                    7: per,
                    8: save,
                    9: ref,
                    10: note,
                })
                new_ws.cell(out_row, 5).number_format = "0"
                new_ws.cell(out_row, 6).number_format = "$#,##0"
                new_ws.cell(out_row, 7).number_format = "$#,##0.00"
                new_ws.cell(out_row, 8).number_format = "0.0%"
                new_ws.cell(out_row, 9).number_format = "$#,##0.00"
                out_row += 1

            changed.append((product, base, p5, p10, p20))
            row += 3
            continue

        write_row_from(row, out_row, {col: ws.cell(row, col).value for col in range(1, 10)})
        out_row += 1
        row += 1

    for merged in ws.merged_cells.ranges:
        if merged.min_row <= 3:
            new_ws.merge_cells(
                start_row=merged.min_row,
                start_column=merged.min_col,
                end_row=merged.max_row,
                end_column=merged.max_col,
            )

    new_wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Products processed: {product_count}")
    print("Sample optimized tiers:")
    for product, base, p5, p10, p20 in changed[:8]:
        print(
            f"- {product}: 1=${base:.2f}; "
            f"5=${p5:.2f}/vial (${p5 * 5:.0f}); "
            f"10=${p10:.2f}/vial (${p10 * 10:.0f}); "
            f"20=${p20:.2f}/vial (${p20 * 20:.0f})"
        )


if __name__ == "__main__":
    main()
