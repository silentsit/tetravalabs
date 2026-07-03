from copy import copy
from pathlib import Path
import re

import openpyxl
from openpyxl.styles import PatternFill


SUPPLIER_PATH = Path(r"C:\Users\user\Downloads\产品价格表.xlsx")
COST_PATH = Path(r"C:\Users\user\Downloads\Peptide Cost Sheet.xlsx")
OUT_PATH = Path(r"C:\Users\user\Downloads\Peptide Cost Sheet.tallied-corrected.xlsx")


# Safe manual aliases where the shop naming differs from supplier naming, but the product/strength matches.
ALIASES = {
    "cagrilintidesemaglutide5mg5mg": "cagrilintidesemaglutide5mg",
    "cagrilintidesemaglutide10mg10mg": "cagrilintidesemaglutide10mg",
    "bpc1575mgtb5005mg10mg": "bpc5mgtb5mg10mg",
    "bpc1575mgtb5005mg20mg": "bpc5mgtb5mg20mg",
    "cu50mgtb50010mgbpc15710mgkpv10mg80mg": "cu50tb10bc10kpv1080mg",
    "glowtb50010mgbpc15710mgghkcu50mg70mg": "glowtb10mgbpc15710mgghk50mg70mg",
    "cjc1295withoutdacipamorelinblend10mg": "cjc1295ipamorelin10mg",
    "cjc1295withoutdacsermorelinipamorelinblend5mg": "cjc1295withoutdacsermorelinipa5mg5mg",
}


def normalize(value):
    text = str(value or "").lower().replace("µ", "u")
    return re.sub(r"[^a-z0-9]+", "", text)


def english_name(value):
    return str(value or "").split("\n")[0].strip()


def copy_cell_style(src_cell, dst_cell):
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)


def load_supplier_records():
    wb = openpyxl.load_workbook(SUPPLIER_PATH, data_only=True)
    ws = wb.active
    records = {}
    last_name = None

    for row in range(3, min(ws.max_row, 1000) + 1):
        col_a = ws.cell(row, 1).value
        col_b = ws.cell(row, 2).value
        spec = ws.cell(row, 3).value
        cost = ws.cell(row, 5).value

        if all(value is None for value in (col_a, col_b, spec, cost)):
            continue

        code = col_a
        name = col_b
        if isinstance(col_a, str) and ("\n" in col_a or any(ch in col_a for ch in "水醋")) and isinstance(col_b, str):
            name = col_a
            code = col_b

        if name:
            last_name = name
        else:
            name = last_name

        if not name or not spec or cost is None:
            continue

        product = english_name(name)
        key = normalize(f"{product} {str(spec).strip()}")
        records[key] = {
            "product": product,
            "spec": str(spec).strip(),
            "cost": cost,
            "row": row,
            "code": code,
        }

    return records


def numeric_equal(left, right):
    try:
        return abs(float(left) - float(right)) < 0.000001
    except (TypeError, ValueError):
        return False


def main():
    supplier = load_supplier_records()
    wb = openpyxl.load_workbook(COST_PATH)
    ws = wb.active

    audit = wb.create_sheet("Cost Correction Audit")
    headers = [
        "Row",
        "Product",
        "Old RMB",
        "Supplier RMB",
        "Action",
        "Match Type",
        "Supplier Product",
        "Supplier Spec",
        "Supplier Row",
    ]
    audit.append(headers)
    for cell in audit[1]:
        copy_cell_style(ws["A1"], cell)

    corrected_fill = PatternFill("solid", fgColor="C6EFCE")
    missing_fill = PatternFill("solid", fgColor="FFF2CC")
    corrections = 0
    exact_checked = 0
    alias_checked = 0
    missing_exact = 0
    already_correct = 0

    for row in range(2, ws.max_row + 1):
        product = ws.cell(row, 2).value
        current = ws.cell(row, 3).value
        if not product:
            continue

        raw_key = normalize(product)
        key = ALIASES.get(raw_key, raw_key)
        match_type = "alias" if key != raw_key else "exact"
        match = supplier.get(key)

        if not match:
            if current is None or current == "":
                ws.cell(row, 3).fill = missing_fill
                missing_exact += 1
                audit.append([row, product, current, None, "MISSING_NO_EXACT_MATCH", "", "", "", ""])
            continue

        if match_type == "alias":
            alias_checked += 1
        else:
            exact_checked += 1

        supplier_cost = match["cost"]
        if current is None or current == "" or str(current).startswith("="):
            ws.cell(row, 3).value = supplier_cost
            ws.cell(row, 3).fill = corrected_fill
            corrections += 1
            action = "FILLED_FROM_SUPPLIER"
        elif numeric_equal(current, supplier_cost):
            already_correct += 1
            action = "ALREADY_CORRECT"
        else:
            ws.cell(row, 3).value = supplier_cost
            ws.cell(row, 3).fill = corrected_fill
            corrections += 1
            action = "CORRECTED_FROM_SUPPLIER"

        audit.append([
            row,
            product,
            current,
            supplier_cost,
            action,
            match_type,
            match["product"],
            match["spec"],
            match["row"],
        ])

    for col in range(1, len(headers) + 1):
        audit.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 24

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    print(f"Exact supplier matches checked: {exact_checked}")
    print(f"Alias supplier matches checked: {alias_checked}")
    print(f"Already correct: {already_correct}")
    print(f"Corrected or filled: {corrections}")
    print(f"Missing with no exact supplier match: {missing_exact}")


if __name__ == "__main__":
    main()
