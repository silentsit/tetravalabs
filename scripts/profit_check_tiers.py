from copy import copy
from pathlib import Path
import re

import openpyxl
from openpyxl.styles import Font, PatternFill


COST_PATH = Path(r"C:\Users\user\Downloads\Peptide Cost Sheet.tallied-corrected.xlsx")
PRICING_PATH = Path(r"C:\Users\user\Downloads\Tiered Pricing-upload.cleaned-with-new-items.xlsx")
OUT_PATH = Path(r"C:\Users\user\Downloads\Tiered Pricing Profit Analysis.xlsx")


# Safe aliases between cost-sheet names and tiered-pricing names.
ALIASES = {
    "cagrilintidesemaglutide5mg5mg": "cagrilintidesemaglutide5mg",
    "cagrilintidesemaglutide10mg10mg": "cagrilintidesemaglutide10mg",
    "bpc1575mgtb5005mg10mg": "bpc1575mgtb5005mg10mg",
    "bpc1575mgtb5005mg20mg": "bpc1575mgtb5005mg20mg",
    "cu50mgtb50010mgbpc15710mgkpv10mg80mg": "cu50mgtb50010mgbpc15710mgkpv10mg80mg",
    "glowtb50010mgbpc15710mgghkcu50mg70mg": "glowtb50010mgbpc15710mgghkcu50mg70mg",
}


def normalize(value):
    text = str(value or "").lower().replace("µ", "u")
    return re.sub(r"[^a-z0-9]+", "", text)


def clone_style(src, dst):
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def extract_fallback_usd(formula):
    if not isinstance(formula, str):
        return None
    # Formula shape: =IFERROR(...,12.345)
    match = re.search(r",\s*(-?\d+(?:\.\d+)?)\s*\)\s*$", formula)
    if not match:
        return None
    return float(match.group(1))


def implied_exchange_rate(ws_values, ws_formulas):
    rates = []
    for row in range(2, ws_values.max_row + 1):
        rmb = ws_values.cell(row, 3).value
        if not isinstance(rmb, (int, float)) or rmb == 0:
            continue
        usd_value = ws_values.cell(row, 4).value
        if isinstance(usd_value, (int, float)):
            rates.append(float(usd_value) / float(rmb))
            continue
        fallback = extract_fallback_usd(ws_formulas.cell(row, 4).value)
        if fallback is not None:
            rates.append(float(fallback) / float(rmb))
    if not rates:
        raise RuntimeError("Could not infer CNY→USD exchange rate from column D.")
    return sum(rates) / len(rates)


def load_costs():
    wb_values = openpyxl.load_workbook(COST_PATH, data_only=True)
    wb_formulas = openpyxl.load_workbook(COST_PATH, data_only=False)
    ws_values = wb_values["Sheet1"]
    ws_formulas = wb_formulas["Sheet1"]
    rate = implied_exchange_rate(ws_values, ws_formulas)

    costs = {}
    missing_cost_rows = []
    for row in range(2, ws_values.max_row + 1):
        product = ws_values.cell(row, 2).value
        rmb = ws_values.cell(row, 3).value
        if not product:
            continue
        key = normalize(product)
        key = ALIASES.get(key, key)
        if isinstance(rmb, (int, float)):
            usd = float(rmb) * rate
            costs[key] = {
                "product": product,
                "rmb": float(rmb),
                "usd": usd,
                "row": row,
            }
        else:
            missing_cost_rows.append((row, product))

    return costs, missing_cost_rows, rate


def main():
    costs, missing_cost_rows, rate = load_costs()

    wb_price = openpyxl.load_workbook(PRICING_PATH, data_only=True)
    ws_price = wb_price["Tiered Pricing"]

    out_wb = openpyxl.Workbook()
    ws = out_wb.active
    ws.title = "Profit Analysis"
    missing_ws = out_wb.create_sheet("Missing Cost Matches")
    summary_ws = out_wb.create_sheet("Summary")

    headers = [
        "Category",
        "Product",
        "Tier",
        "Qty",
        "Sell Total USD",
        "Sell Per Vial USD",
        "Cost Per Vial USD",
        "Profit Per Vial USD",
        "Total Profit USD",
        "Margin %",
        "Status",
        "Cost Sheet Row",
        "Cost RMB",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    loss_fill = PatternFill("solid", fgColor="F4CCCC")
    missing_fill = PatternFill("solid", fgColor="FFF2CC")
    ok_fill = PatternFill("solid", fgColor="D9EAD3")

    offer_rows = 0
    profitable_rows = 0
    loss_rows = 0
    missing_rows = 0
    products_with_missing = set()
    products_with_loss = set()
    min_profit_row = None
    total_profit_5 = total_profit_10 = total_profit_20 = 0.0

    for row in range(1, ws_price.max_row + 1):
        category = ws_price.cell(row, 2).value
        product = ws_price.cell(row, 3).value
        tier = ws_price.cell(row, 4).value
        qty = ws_price.cell(row, 5).value
        total_price = ws_price.cell(row, 6).value
        per_vial = ws_price.cell(row, 7).value

        if not product or qty not in (5, 10, 20):
            continue

        offer_rows += 1
        key = normalize(product)
        cost = costs.get(key)
        if not cost:
            status = "MISSING COST"
            out = [category, product, tier, qty, total_price, per_vial, None, None, None, None, status, None, None]
            ws.append(out)
            for cell in ws[ws.max_row]:
                cell.fill = missing_fill
            missing_rows += 1
            products_with_missing.add(product)
            continue

        cost_usd = cost["usd"]
        profit_per_vial = float(per_vial) - cost_usd
        total_profit = profit_per_vial * int(qty)
        margin = profit_per_vial / float(per_vial) if per_vial else None
        status = "LOSS" if profit_per_vial < 0 else "PROFIT"
        if status == "LOSS":
            loss_rows += 1
            products_with_loss.add(product)
        else:
            profitable_rows += 1
        if qty == 5:
            total_profit_5 += total_profit
        elif qty == 10:
            total_profit_10 += total_profit
        elif qty == 20:
            total_profit_20 += total_profit
        if min_profit_row is None or profit_per_vial < min_profit_row["profit_per_vial"]:
            min_profit_row = {
                "product": product,
                "tier": tier,
                "profit_per_vial": profit_per_vial,
                "margin": margin,
            }

        ws.append([
            category,
            product,
            tier,
            qty,
            total_price,
            per_vial,
            cost_usd,
            profit_per_vial,
            total_profit,
            margin,
            status,
            cost["row"],
            cost["rmb"],
        ])
        fill = loss_fill if status == "LOSS" else ok_fill
        ws.cell(ws.max_row, 11).fill = fill

    for col in (5, 6, 7, 8, 9):
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
            for c in cell:
                c.number_format = "$#,##0.00"
    for c in ws.iter_cols(min_col=10, max_col=10, min_row=2):
        for cell in c:
            cell.number_format = "0.0%"
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    missing_ws.append(["Pricing Product Missing Cost", "Pricing Rows Affected"])
    for cell in missing_ws[1]:
        cell.font = Font(bold=True)
        cell.fill = missing_fill
    for product in sorted(products_with_missing):
        missing_ws.append([product, 3])
    missing_ws.append([])
    missing_ws.append(["Cost Sheet Rows With Blank RMB/USD"])
    for row, product in missing_cost_rows:
        missing_ws.append([row, product])
    for col in range(1, 4):
        missing_ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 35

    summary_rows = [
        ("Implied CNY→USD rate used", rate),
        ("Offer rows checked", offer_rows),
        ("Profitable rows", profitable_rows),
        ("Loss rows", loss_rows),
        ("Rows missing cost", missing_rows),
        ("Products missing cost", len(products_with_missing)),
        ("Total profit if selling one 5-vial pack of each costed item", total_profit_5),
        ("Total profit if selling one 10-vial pack of each costed item", total_profit_10),
        ("Total profit if selling one 20-vial pack of each costed item", total_profit_20),
    ]
    if min_profit_row:
        summary_rows.extend([
            ("Lowest profit item/tier", f"{min_profit_row['product']} - {min_profit_row['tier']}"),
            ("Lowest profit per vial", min_profit_row["profit_per_vial"]),
            ("Lowest margin", min_profit_row["margin"]),
        ])
    summary_ws.append(["Metric", "Value"])
    for row in summary_rows:
        summary_ws.append(list(row))
    for cell in summary_ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row in range(2, summary_ws.max_row + 1):
        if isinstance(summary_ws.cell(row, 2).value, float):
            summary_ws.cell(row, 2).number_format = "$#,##0.00" if "profit" in str(summary_ws.cell(row, 1).value).lower() else "0.0000"
    summary_ws.column_dimensions["A"].width = 58
    summary_ws.column_dimensions["B"].width = 28

    out_wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    print(f"Implied CNY→USD rate: {rate:.8f}")
    print(f"Offer rows checked: {offer_rows}")
    print(f"Profitable rows: {profitable_rows}")
    print(f"Loss rows: {loss_rows}")
    print(f"Rows missing cost: {missing_rows}")
    print(f"Products missing cost: {len(products_with_missing)}")
    if min_profit_row:
        print(
            "Lowest profit: "
            f"{min_profit_row['product']} / {min_profit_row['tier']} = "
            f"${min_profit_row['profit_per_vial']:.2f} per vial "
            f"({min_profit_row['margin']:.1%} margin)"
        )


if __name__ == "__main__":
    main()
