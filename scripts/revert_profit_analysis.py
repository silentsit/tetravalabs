from pathlib import Path
import re

import openpyxl
from openpyxl.styles import Font, PatternFill

COST_PATH = Path(r"C:\Users\user\Downloads\Peptide Cost Sheet.tallied-corrected.xlsx")
PRICING_PATH = Path(r"C:\Users\user\Downloads\Tiered Pricing-upload.cleaned-with-new-items.xlsx")
OUT_PATH = Path(r"C:\Users\user\Downloads\Tiered Pricing Profit Analysis.static.xlsx")

ALIASES = {
    "cagrilintidesemaglutide5mg5mg": "cagrilintidesemaglutide5mg",
    "cagrilintidesemaglutide10mg10mg": "cagrilintidesemaglutide10mg",
}


def normalize(value):
    text = str(value or "").lower().replace("µ", "u")
    return re.sub(r"[^a-z0-9]+", "", text)


def extract_fallback_usd(formula):
    if not isinstance(formula, str):
        return None
    match = re.search(r",\s*(-?\d+(?:\.\d+)?)\s*\)\s*$", formula)
    return float(match.group(1)) if match else None


def implied_exchange_rate(ws_values, ws_formulas):
    rates = []
    for row in range(2, ws_values.max_row + 1):
        rmb = ws_values.cell(row, 3).value
        if not isinstance(rmb, (int, float)) or rmb == 0:
            continue
        usd_value = ws_values.cell(row, 4).value
        if isinstance(usd_value, (int, float)):
            rates.append(float(usd_value) / float(rmb))
            continue
        fallback = extract_fallback_usd(ws_formulas.cell(row, 4).value)
        if fallback is not None:
            rates.append(float(fallback) / float(rmb))
    return sum(rates) / len(rates)


def load_costs():
    wb_values = openpyxl.load_workbook(COST_PATH, data_only=True)
    wb_formulas = openpyxl.load_workbook(COST_PATH, data_only=False)
    ws_values = wb_values["Sheet1"]
    ws_formulas = wb_formulas["Sheet1"]
    rate = implied_exchange_rate(ws_values, ws_formulas)
    costs = {}
    missing_cost_rows = []
    for row in range(2, ws_values.max_row + 1):
        product = ws_values.cell(row, 2).value
        rmb = ws_values.cell(row, 3).value
        if not product:
            continue
        key = ALIASES.get(normalize(product), normalize(product))
        if isinstance(rmb, (int, float)):
            costs[key] = {"product": product, "rmb": float(rmb), "usd": float(rmb) * rate, "row": row}
        else:
            missing_cost_rows.append((row, product))
    return costs, missing_cost_rows, rate


def load_pricing_groups():
    wb = openpyxl.load_workbook(PRICING_PATH, data_only=True)
    ws = wb["Tiered Pricing"]
    by_product = {}

    for row in range(1, ws.max_row + 1):
        category = ws.cell(row, 2).value
        product = ws.cell(row, 3).value
        qty = ws.cell(row, 5).value
        total = ws.cell(row, 6).value
        per_vial = ws.cell(row, 7).value
        if not product or qty not in (1, 5, 10, 20):
            continue

        group = by_product.setdefault(product, {
            "category": category,
            "product": product,
            "ref_total": None,
            "ref_per_vial": None,
            "tiers": [],
        })
        if category:
            group["category"] = category
        if qty == 1:
            group["ref_total"] = total
            group["ref_per_vial"] = per_vial
        else:
            group["tiers"].append({
                "tier": ws.cell(row, 4).value,
                "qty": int(qty),
                "total": total,
                "per_vial": per_vial,
            })

    return list(by_product.values())


def main():
    costs, missing_cost_rows, rate = load_costs()
    groups = load_pricing_groups()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profit Analysis"
    missing_ws = wb.create_sheet("Missing Cost Matches")
    summary_ws = wb.create_sheet("Summary")

    headers = [
        "Category", "Product", "Tier", "Qty", "Sell Total USD", "Sell Per Vial USD",
        "Cost Per Vial USD", "Profit Per Vial USD", "Total Profit USD", "Margin %",
        "Status", "Cost Sheet Row", "Cost RMB",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    ref_fill = PatternFill("solid", fgColor="E8EEF7")
    ok_fill = PatternFill("solid", fgColor="D9EAD3")
    loss_fill = PatternFill("solid", fgColor="F4CCCC")
    missing_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    offer_rows = ref_rows = profitable_rows = loss_rows = missing_rows = 0
    products_with_missing = set()
    min_profit_row = None

    def append_row(values, fill=None):
        ws.append(values)
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill

    for group in groups:
        product = group["product"]
        cost = costs.get(normalize(product))
        cost_usd = cost["usd"] if cost else None
        cost_row = cost["row"] if cost else None
        cost_rmb = cost["rmb"] if cost else None

        if group["ref_per_vial"] is not None:
            ref_per = float(group["ref_per_vial"])
            ref_total = float(group["ref_total"]) if group["ref_total"] is not None else ref_per
            ref_profit = ref_per - cost_usd if cost_usd is not None else None
            ref_margin = ref_profit / ref_per if ref_profit is not None and ref_per else None
            append_row([
                group["category"], product, "1 Vial (Ref)", 1, ref_total, ref_per,
                cost_usd, ref_profit, ref_profit, ref_margin, "REFERENCE", cost_row, cost_rmb,
            ], ref_fill)
            ref_rows += 1

        for tier in sorted(group["tiers"], key=lambda t: t["qty"]):
            offer_rows += 1
            per_vial = float(tier["per_vial"])
            qty = tier["qty"]
            total = float(tier["total"])
            if cost_usd is None:
                append_row([
                    group["category"], product, tier["tier"], qty, total, per_vial,
                    None, None, None, None, "MISSING COST", None, None,
                ], missing_fill)
                missing_rows += 1
                products_with_missing.add(product)
                continue
            profit_per_vial = per_vial - cost_usd
            total_profit = profit_per_vial * qty
            margin = profit_per_vial / per_vial if per_vial else None
            status = "LOSS" if profit_per_vial < 0 else "PROFIT"
            fill = loss_fill if status == "LOSS" else ok_fill
            if status == "LOSS":
                loss_rows += 1
            else:
                profitable_rows += 1
            if min_profit_row is None or profit_per_vial < min_profit_row["profit_per_vial"]:
                min_profit_row = {"product": product, "tier": tier["tier"], "profit_per_vial": profit_per_vial, "margin": margin}
            append_row([
                group["category"], product, tier["tier"], qty, total, per_vial,
                cost_usd, profit_per_vial, total_profit, margin, status, cost_row, cost_rmb,
            ], fill)

    for col in (5, 6, 7, 8, 9):
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
            for c in cell:
                c.number_format = "$#,##0.00"
    for cell in ws.iter_cols(min_col=10, max_col=10, min_row=2):
        for c in cell:
            c.number_format = "0.0%"
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    missing_ws.append(["Pricing Product Missing Cost", "Offer Rows Affected"])
    for product in sorted(products_with_missing):
        missing_ws.append([product, 3])
    missing_ws.append([])
    missing_ws.append(["Cost Sheet Rows With Blank RMB/USD"])
    for row, product in missing_cost_rows:
        missing_ws.append([row, product])

    summary_ws.append(["Metric", "Value"])
    summary_rows = [
        ("Implied CNY→USD rate used", rate),
        ("Reference rows (1 vial)", ref_rows),
        ("Offer rows checked (5/10/20)", offer_rows),
        ("Profitable offer rows", profitable_rows),
        ("Loss offer rows", loss_rows),
        ("Offer rows missing cost", missing_rows),
        ("Products missing cost", len(products_with_missing)),
    ]
    if min_profit_row:
        summary_rows.extend([
            ("Lowest offer-tier profit item", f"{min_profit_row['product']} - {min_profit_row['tier']}"),
            ("Lowest offer-tier profit per vial", min_profit_row["profit_per_vial"]),
            ("Lowest offer-tier margin", min_profit_row["margin"]),
        ])
    for row in summary_rows:
        summary_ws.append(list(row))

    wb.save(OUT_PATH)
    print(f"Reverted {OUT_PATH}")
    print(f"Reference rows: {ref_rows}, Offer rows: {offer_rows}")


if __name__ == "__main__":
    main()
