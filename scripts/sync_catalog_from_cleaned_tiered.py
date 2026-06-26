"""Sync product_catalog_usd.json and tiered-catalog.json from cleaned tiered pricing."""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
TIERED_DEFAULT = Path.home() / "Downloads" / "Tiered Pricing-upload.cleaned-with-new-items.xlsx"
CATALOG_PATH = ROOT / "product_catalog_usd.json"
TIERED_PATH = ROOT / "packages" / "catalog" / "data" / "tiered-catalog.json"

SOURCE_CATEGORY_MAP = {
    "Supplies & Reconstitution": "Lab Supplies",
    "GLP-1 / Incretin": "GLP-1 Research",
    "BPC-157 / TB500": "Growth Factors",
    "Blends": "Research Blends",
    "CJC / Ipamorelin / GHRP": "Growth Factors",
    "Growth Hormone Axis": "Growth Factors",
    "Mitochondrial / Metabolic Other": "Growth Factors",
    "Cosmetic / Copper / Tanning": "Growth Factors",
    "Longevity / Thymic / Neuropeptides": "Growth Factors",
    "Vitamins & Injectables": "Growth Factors",
    "Legacy Catalog": "Growth Factors",
}

BLEND_PRODUCTS = {
    "BPC-157 5mg + TB500 5mg",
    "CU 50mg + TB500 10mg + BPC-157 10mg + KPV 10mg",
    "Glow BPC-157 + TB500 + GHK-Cu",
    "Glow TB500 10mg + BPC-157 10mg + GHK-Cu 50mg",
    "CJC-1295 without DAC / Ipamorelin Blend",
    "CJC-1295 without DAC / Sermorelin / Ipamorelin Blend",
    "Cagrilintide + Semaglutide",
}
GROWTH_FROM_GLP1 = {"Sermorelin", "Tesamorelin"}
GLP1_PRODUCTS = {
    "Semaglutide",
    "Tirzepatide",
    "Retatrutide",
    "Cagrilintide",
    "Mazdutide",
    "Survodutide",
    "5-Amino-1MQ",
    "AOD-9604",
}


def slugify(value: str) -> str:
    return re.sub(r"^-+|-+$", "", re.sub(r"[^a-z0-9]+", "-", value.lower()))


def parse_strength(name: str) -> tuple[str, str]:
    match = re.search(r"(\d+\s*(?:mg|ml|iu|mcg|count(?:\s*\([^)]+\))?)[^\s]*)", name, re.I)
    if not match:
        return name.strip(), "Standard"
    strength = match.group(1).strip()
    base_name = re.sub(r"\s+" + re.escape(strength) + r"$", "", name, flags=re.I).strip()
    return base_name, strength


def resolve_storefront_category(base_name: str, source_category: str) -> str:
    if base_name in BLEND_PRODUCTS:
        return "Research Blends"
    if source_category == "GLP-1 / Incretin":
        if base_name in GROWTH_FROM_GLP1:
            return "Growth Factors"
        if base_name in GLP1_PRODUCTS:
            return "GLP-1 Research"
        return "Growth Factors"
    return SOURCE_CATEGORY_MAP.get(source_category, "Growth Factors")


def load_tiered_products(xlsx_path: Path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Tiered Pricing"]
    products: dict[str, dict] = {}
    order: list[str] = []
    last_category = None

    for row in range(4, ws.max_row + 1):
        category = ws.cell(row, 2).value
        name = ws.cell(row, 3).value
        tier = ws.cell(row, 4).value
        qty = ws.cell(row, 5).value
        total = ws.cell(row, 6).value
        per = ws.cell(row, 7).value
        savings = ws.cell(row, 8).value

        if category and not name:
            last_category = str(category).strip()
            continue
        if not name or name in ("Product Name", "Category"):
            continue
        if tier in ("Tier", "Pricing Note") or qty in (None, "Qty"):
            continue

        if category:
            last_category = str(category).strip()

        entry = products.setdefault(
            name,
            {"category": last_category, "name": name, "packs": {}},
        )
        if qty in (5, 10, 20):
            entry["packs"][int(qty)] = {
                "total": round(float(total), 2),
                "per": round(float(per), 2),
                "save": round(float(savings), 4) if savings is not None else 0,
            }
        if name not in order:
            order.append(name)

    return products, order


def match_existing_slug(display_name: str, old_rows: list[dict]) -> tuple[str, str, str, str | None]:
    normalized = display_name.strip().lower()
    for row in old_rows:
        combined = f"{row['name']} {row['strength']}".strip().lower()
        if combined == normalized or row["name"].strip().lower() == normalized:
            return row["slug"], row["name"], row["strength"], row.get("storefront_category")
    base_name, strength = parse_strength(display_name)
    return slugify(display_name), base_name, strength, None


def build_rows(products: dict, order: list[str], old_rows: list[dict], source_path: Path):
    tiered_products = []
    flat_rows = []

    for display_name in order:
        data = products[display_name]
        source_category = (data["category"] or "Legacy Catalog").strip()
        slug, base_name, strength, existing_storefront = match_existing_slug(display_name, old_rows)
        storefront_category = existing_storefront or resolve_storefront_category(base_name, source_category)

        pack_tiers = []
        for qty in (5, 10, 20):
            pack = data["packs"].get(qty)
            if not pack:
                continue
            pack_tiers.append(
                {
                    "tier": f"{qty} vials",
                    "qty": qty,
                    "price_usd": pack["total"],
                    "per_unit_usd": pack["per"],
                    "savings_pct": pack["save"],
                }
            )

        if not pack_tiers:
            continue

        tiered_products.append(
            {
                "category": source_category,
                "storefront_category": storefront_category,
                "name": display_name,
                "slug": slug,
                "pack_tiers": pack_tiers,
            }
        )
        flat_rows.append(
            {
                "category": source_category,
                "name": base_name,
                "strength": strength,
                "price_usd": pack_tiers[0]["price_usd"],
                "slug": slug,
                "storefront_category": storefront_category,
                "pack_tiers": pack_tiers,
            }
        )

    return {"source": str(source_path), "products": tiered_products}, flat_rows


def main() -> int:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else TIERED_DEFAULT
    if not xlsx_path.exists():
        print(f"Missing pricing workbook: {xlsx_path}", file=sys.stderr)
        return 1

    old_rows = json.loads(CATALOG_PATH.read_text(encoding="utf-8")) if CATALOG_PATH.exists() else []
    products, order = load_tiered_products(xlsx_path)
    tiered_payload, flat_rows = build_rows(products, order, old_rows, xlsx_path)

    CATALOG_PATH.write_text(json.dumps(flat_rows, indent=2), encoding="utf-8")
    TIERED_PATH.write_text(json.dumps(tiered_payload, indent=2), encoding="utf-8")

    print(f"Synced {len(flat_rows)} products from {xlsx_path}")
    print(f"Updated {CATALOG_PATH}")
    print(f"Updated {TIERED_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
