"""Update Price List USD.xlsx from cleaned tiered pricing + catalog SKUs."""

from __future__ import annotations

import json
import re
import shutil
from copy import copy
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[1]
DOWNLOADS = Path.home() / "Downloads"
TIERED_PATH = DOWNLOADS / "Tiered Pricing-upload.cleaned-with-new-items.xlsx"
OUTPUT_PATH = DOWNLOADS / "Price List USD.xlsx"
TEMPLATE_PATH = DOWNLOADS / "Customer Tiered Price List USD.xlsx"
CATALOG_JSON = ROOT / "product_catalog_usd.json"
NORMALIZED_JSON = ROOT / "packages" / "catalog" / "output" / "catalog.normalized.json"

NEW_PRODUCTS = [
    "Pinealon 10mg",
    "Humanin 10mg",
    "Adamax 10mg",
    "Dihexa 10mg",
    "Cerebrolysin 10mg",
]


def slugify(value: str) -> str:
    return re.sub(r"^-+|-+$", "", re.sub(r"[^a-z0-9]+", "-", value.lower()))


def product_sku(slug: str) -> str:
    return slug.replace("-", "_").upper()


def pack_sku(slug: str, qty: int) -> str:
    return f"{product_sku(slug)}_{qty}PK"


def clone_style(src, dst):
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def load_tiered_products():
    wb = openpyxl.load_workbook(TIERED_PATH, data_only=True)
    ws = wb["Tiered Pricing"]
    products: dict[str, dict] = {}
    order: list[str] = []
    last_category = None

    for row in range(4, ws.max_row + 1):
        category = ws.cell(row, 2).value
        name = ws.cell(row, 3).value
        tier = ws.cell(row, 4).value
        qty = ws.cell(row, 5).value
        total = ws.cell(row, 6).value
        per = ws.cell(row, 7).value
        savings = ws.cell(row, 8).value

        if category and not name:
            last_category = str(category).strip()
            continue
        if not name or name in ("Product Name", "Category"):
            continue
        if tier in ("Tier", "Pricing Note") or qty in (None, "Qty"):
            continue

        if category:
            last_category = str(category).strip()

        entry = products.setdefault(
            name,
            {"category": last_category, "name": name, "ref": None, "packs": {}},
        )
        if qty == 1:
            entry["ref"] = float(per) if per is not None else None
        elif qty in (5, 10, 20):
            entry["packs"][int(qty)] = {
                "total": float(total),
                "per": float(per),
                "save": float(savings) if savings is not None else 0,
            }
        if name not in order:
            order.append(name)

    return products, order


def load_slug_and_sku_maps():
    slug_by_name: dict[str, str] = {}
    sku_by_name: dict[str, str] = {}
    pack_skus_by_name: dict[str, dict[int, str]] = {}

    if CATALOG_JSON.exists():
        for row in json.loads(CATALOG_JSON.read_text(encoding="utf-8")):
            display = f"{row['name']} {row['strength']}".strip()
            slug_by_name[display] = row["slug"]
            sku_by_name[display] = product_sku(row["slug"])

    if NORMALIZED_JSON.exists():
        catalog = json.loads(NORMALIZED_JSON.read_text(encoding="utf-8"))
        for product in catalog.get("products", []):
            title = product["title"].replace(" 100 count)", "").strip()
            slug = product["handle"]
            slug_by_name[title] = slug
            sku_by_name[title] = product_sku(slug)
            packs: dict[int, str] = {}
            for variant in product.get("variants", []):
                qty = variant.get("metadata", {}).get("pack_qty")
                if qty in (5, 10, 20):
                    packs[int(qty)] = variant["sku"]
            if packs:
                pack_skus_by_name[title] = packs

    for name in NEW_PRODUCTS:
        slug = slugify(name)
        slug_by_name[name] = slug
        sku_by_name[name] = product_sku(slug)

    return slug_by_name, sku_by_name, pack_skus_by_name


def resolve_slug(name: str, slug_by_name: dict[str, str]) -> str:
    return slug_by_name.get(name, slugify(name))


def detect_layout(ws):
    for row in range(1, 8):
        if ws.cell(row, 2).value == "Category" or ws.cell(row, 3).value == "Product Name":
            return row
    return 4


def unmerge_data_area(ws, start_row: int):
    to_remove = []
    for merged in ws.merged_cells.ranges:
        if merged.min_row >= start_row:
            to_remove.append(str(merged))
    for ref in to_remove:
        ws.unmerge_cells(ref)


def write_product_row(ws, row, template_row, data, slug, sku_by_name, pack_skus_by_name):
    unmerge_data_area(ws, row)

    for col in range(1, 17):
        clone_style(ws.cell(template_row, max(col, 2)), ws.cell(row, col))

    name = data["name"]
    base_sku = sku_by_name.get(name, product_sku(slug))
    pack_map = pack_skus_by_name.get(name, {})

    ws.cell(row, 1).value = base_sku
    ws.cell(row, 2).value = None
    ws.cell(row, 3).value = name
    ws.cell(row, 4).value = data["ref"]
    for qty, cols in [(5, (5, 6, 7)), (10, (8, 9, 10)), (20, (11, 12, 13))]:
        pack = data["packs"].get(qty)
        if not pack:
            continue
        ws.cell(row, cols[0]).value = pack["total"]
        ws.cell(row, cols[1]).value = pack["per"]
        ws.cell(row, cols[2]).value = pack["save"]

    ws.cell(row, 14).value = pack_map.get(5, pack_sku(slug, 5))
    ws.cell(row, 15).value = pack_map.get(10, pack_sku(slug, 10))
    ws.cell(row, 16).value = pack_map.get(20, pack_sku(slug, 20))

    ws.cell(row, 4).number_format = "$#,##0.00"
    for col in (5, 6, 8, 9, 11, 12):
        ws.cell(row, col).number_format = "$#,##0.00"
    for col in (7, 10, 13):
        ws.cell(row, col).number_format = "0.0%"


def main():
    if not TIERED_PATH.exists():
        raise FileNotFoundError(f"Missing pricing source: {TIERED_PATH}")

    source_template = OUTPUT_PATH if OUTPUT_PATH.exists() else TEMPLATE_PATH
    if not source_template.exists():
        raise FileNotFoundError(f"Missing template workbook: {source_template}")

    if source_template != OUTPUT_PATH:
        shutil.copy2(source_template, OUTPUT_PATH)

    products, tier_order = load_tiered_products()
    slug_by_name, sku_by_name, pack_skus_by_name = load_slug_and_sku_maps()

    wb = openpyxl.load_workbook(OUTPUT_PATH)
    ws = wb.active
    header_row = detect_layout(ws)
    data_start = header_row + 1

    today = date.today().strftime("%d/%m/%Y")
    ws["B2"] = f"PRICE LIST USD [{today}]"

    ws.cell(header_row, 1).value = "SKU"
    ws.cell(header_row, 14).value = "SKU 5-Pack"
    ws.cell(header_row, 15).value = "SKU 10-Pack"
    ws.cell(header_row, 16).value = "SKU 20-Pack"
    for col in (1, 14, 15, 16):
        clone_style(ws.cell(header_row, 2), ws.cell(header_row, col))
        ws.cell(header_row, col).font = Font(bold=True)

    categories: list[str] = []
    for row in range(data_start, ws.max_row + 1):
        cat = ws.cell(row, 2).value
        prod = ws.cell(row, 3).value
        if cat and not prod and str(cat).strip().lower() != "category":
            categories.append(str(cat).strip())

    cat_products: dict[str, list[str]] = {}
    cat_order: list[str] = []
    for name in tier_order:
        cat = products[name]["category"] or "Uncategorized"
        if cat not in cat_products:
            cat_products[cat] = []
            cat_order.append(cat)
        cat_products[cat].append(name)

    normalized_cats = {c.strip(): c for c in categories}
    template_product_row = data_start + 1
    for row in range(data_start, ws.max_row + 1):
        if ws.cell(row, 3).value:
            template_product_row = row
            break

    if ws.max_row >= data_start:
        unmerge_data_area(ws, data_start)
        ws.delete_rows(data_start, ws.max_row - data_start + 1)

    current_row = data_start
    updated = 0
    for cat in cat_order:
        label = normalized_cats.get(cat.strip(), f"  {cat}")
        ws.cell(current_row, 2).value = label
        clone_style(ws.cell(data_start, 2), ws.cell(current_row, 2))
        try:
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=13)
        except ValueError:
            pass
        current_row += 1

        for name in cat_products[cat]:
            slug = resolve_slug(name, slug_by_name)
            write_product_row(
                ws,
                current_row,
                template_product_row,
                products[name],
                slug,
                sku_by_name,
                pack_skus_by_name,
            )
            current_row += 1
            updated += 1

    ws.column_dimensions["A"].width = 28
    for col in ("N", "O", "P"):
        ws.column_dimensions[col].width = 28

    wb.save(OUTPUT_PATH)
    print(f"Updated {OUTPUT_PATH}")
    print(f"Template used: {source_template.name}")
    print(f"Products written: {updated}")
    print(f"New items: {[n for n in NEW_PRODUCTS if n in products]}")


if __name__ == "__main__":
    main()
