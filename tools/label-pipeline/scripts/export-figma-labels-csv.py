"""Export Figma batch-import CSV + compositing manifest from Peptide Sequence Sheet NEW.xlsx."""
from __future__ import annotations

import csv
import re
from pathlib import Path

import openpyxl
from openpyxl.cell.rich_text import CellRichText

DEFAULT_XLSX = Path(r"c:\Users\user\Downloads\Peptide Sequence Sheet NEW.xlsx")
DATA_DIR = Path(__file__).resolve().parents[1] / "data"
DEFAULT_BATCH_CSV = DATA_DIR / "labels-batch.csv"
DEFAULT_MANIFEST_CSV = DATA_DIR / "labels-manifest.csv"


def cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, CellRichText):
        return "".join(str(block) for block in value)
    return str(value).strip()


def format_cas(value: str) -> str:
    if not value or value in {"—", "-", "N/A"}:
        return ""
    text = value.strip()
    if text.upper().startswith("CAS"):
        return text
    return f"CAS {text}"


def slug_filename(product: str, concentration: str) -> str:
    base = product.strip()
    dose = concentration.strip()
    if dose and dose not in {"—", "-", "N/A"} and dose.lower() not in base.lower():
        base = f"{base} {dose}"
    base = re.sub(r'[<>:"/\\|?*]', "-", base)
    return re.sub(r"\s+", " ", base).strip()


def is_capsule(product: str, vial_size: str) -> bool:
    combined = f"{product} {vial_size}".lower()
    return "capsule" in combined or re.search(r"\b\d+\s*count\b", combined) is not None


def export_labels(
    xlsx_path: Path = DEFAULT_XLSX,
    batch_csv: Path = DEFAULT_BATCH_CSV,
    manifest_csv: Path = DEFAULT_MANIFEST_CSV,
) -> int:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Peptide Formulas"]

    headers = {ws.cell(1, col).value: col for col in range(1, ws.max_column + 1)}
    product_col = headers["#product_name"]
    sub_col = headers.get("#sub_name")
    cas_col = headers.get("#cas_number")
    formula_col = headers.get("#formula")
    concentration_col = headers.get("#concentration")
    vial_col = headers.get("Vial Size")

    rows: list[dict[str, str]] = []
    for row_idx in range(2, ws.max_row + 1):
        product = cell_text(ws.cell(row_idx, product_col).value)
        if not product:
            continue

        sub_name = cell_text(ws.cell(row_idx, sub_col).value) if sub_col else ""
        concentration = cell_text(ws.cell(row_idx, concentration_col).value)
        vial_size = cell_text(ws.cell(row_idx, vial_col).value) if vial_col else ""
        mockup = "capsule" if is_capsule(product, vial_size) else "vial"
        if mockup == "capsule":
            label_template = "capsule"
        elif sub_name:
            label_template = "flower"
        else:
            label_template = "main"
        export_name = slug_filename(product, concentration)

        rows.append(
            {
                "#product_name": product,
                "#sub_name": sub_name,
                "#cas_number": format_cas(cell_text(ws.cell(row_idx, cas_col).value)) if cas_col else "",
                "#formula": cell_text(ws.cell(row_idx, formula_col).value) if formula_col else "",
                "#concentration": concentration if concentration not in {"—", "-"} else "N/A",
                "label_template": label_template,
                "mockup": mockup,
                "export_filename": export_name,
            }
        )

    DATA_DIR.mkdir(parents=True, exist_ok=True)

    batch_fields = [
        "#product_name",
        "#sub_name",
        "#cas_number",
        "#formula",
        "#concentration",
        "label_template",
        "export_filename",
    ]
    with batch_csv.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=batch_fields)
        writer.writeheader()
        writer.writerows([{k: row[k] for k in batch_fields} for row in rows])

    manifest_fields = ["export_filename", "label_template", "mockup", "#product_name", "#concentration"]
    with manifest_csv.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=manifest_fields)
        writer.writeheader()
        writer.writerows([{k: row[k] for k in manifest_fields} for row in rows])

    return len(rows)


def main() -> None:
    count = export_labels()
    print(f"Wrote {count} rows to {DEFAULT_BATCH_CSV}")
    print(f"Wrote manifest to {DEFAULT_MANIFEST_CSV}")


if __name__ == "__main__":
    main()
